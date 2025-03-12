from app.db.database import MongoClient
from app.repositories.company import CompanyRepository
from fastapi import UploadFile
from app.utils.math import safe_float
from datetime import datetime
import io
from openpyxl import load_workbook
from app.utils.id_gen import generate_uuid_v4_without_special_chars


class CompanyExcelUtil:
    def __init__(self, mongo_client: MongoClient, repository: CompanyRepository):
        self.client = mongo_client
        self.repository = repository

    async def parse_imported_data(self, file: UploadFile, companies: dict) -> bool:
        """
        Read in the CSV file and extract the fields. If the Company already exists, just append the new financials.
        If it is a new company, add it to a list to then insert at once.
        :param client: Mongo Client
        :param file: File to upload
        :return:
        """

        f = await file.read()
        xlsx = io.BytesIO(f)
        wb = load_workbook(xlsx)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        s = set()
        for key, value in companies.items():
            s.add(value.legal_name)
        new_companies_to_add = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = {headers[i]: cell for i, cell in enumerate(row)}
            if not row_data.get('Legal name'):
                continue
            financials = {
                'checking_account': safe_float(row_data.get('Encaisse - Comptes bancaires')),
                'long_term_investments': safe_float(row_data.get('Placements long terme')),
                'total_investments': safe_float(row_data.get('Placements totaux')),
                'physical_assets': safe_float(row_data.get('Terrain et immeubles')),
                'total_actives': safe_float(row_data.get('Total actif')),
                'loans': safe_float(row_data.get('Hypothèques ou crédit potentiel')),
                'total_passives': safe_float(row_data.get('Total Passif')),
                'total_donations': safe_float(row_data.get('Total dons')),
                'federal_revenue': safe_float(row_data.get('Revenus fédéraux')),
                'provincial_revenue': safe_float(row_data.get('Revenus provinciaux')),
                'municipal_revenue': safe_float(row_data.get('Revenus municipaux')),
                'interest_and_banking_fees': safe_float(row_data.get('Intérêts et frais bancaires')),
                'occupation_cost': safe_float(row_data.get("Coûts d'occupation")),
                'professional_fees': safe_float(row_data.get('Honoraires professionnels')),
                'salaries': safe_float(row_data.get('Salaires')),
                'fixed_asset_depreciation': safe_float(row_data.get('Amortissement immobilisations')),
                'others': safe_float(row_data.get('Autres')),
                'total_expenses': safe_float(row_data.get('Total dépenses')),
                'total_revenue': safe_float(row_data.get('Total des revenus')),
                'timestamp': datetime.now()
            }
            if row_data.get('Legal name') not in s:
                email = row_data['Contact Email'] if row_data['Contact Email'] is not None and isinstance(row_data['Contact Email'], str) else ''
                company = {
                    'legal_name': row_data.get('Legal name'),
                    "is_existing_client": bool(row_data['IND_BNC']),
                    "is_active": bool(row_data['CLIENT_ACTIF']),
                    'company_phone_number': str(row_data['Contact Phone']) if row_data['Contact Phone'] is not None else '',
                    'company_email': email,
                    'company_website': '',
                    'description': '',
                    'fcc': 1 if row_data['FCC'] is not None else 0,
                    'street_address': row_data['Mailing address'] if row_data['Mailing address'] is not None else '',
                    'city': row_data['City'] if row_data['City'] is not None else '',
                    'state_or_province': row_data['Province'] if row_data['Province'] is not None else '',
                    'postal_code': row_data['Postal code'] if row_data['Postal code'] is not None else '',
                    'country': 'CA',
                    'contacts': [] if 'Contact Email' not in row_data else [
                        {
                            "id": generate_uuid_v4_without_special_chars(),
                            "name": str(row_data['Directeur de compte']) if row_data['Directeur de compte'] is not None else '',
                            "email": email,
                            "phone_number": str(row_data['Contact Phone']) if row_data['Contact Phone'] is not None else ''
                        }
                    ],
                    'actions': [],
                    'comments': [],
                    'news': [],
                    'financials': [financials]
                }
                new_companies_to_add.append(company)
            else:
                await self.repository.update(self.client, {"$push": {"financials": financials}}, legal_name=row_data.get('Legal name'))
        if len(new_companies_to_add) > 0:
            await self.repository.create_many(self.client, new_companies_to_add)
        return True
